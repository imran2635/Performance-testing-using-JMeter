import openpyxl


def row_count(file, sheet):
    workbook = openpyxl.load_workbook(file)
    worksheet = workbook.get_sheet_by_name(sheet)
    return worksheet.max_row


def column_count(file, sheet):
    workbook = openpyxl.load_workbook(file)
    worksheet = workbook.get_sheet_by_name(sheet)
    return worksheet.max_column


def read_data(file, sheet, reading_row, reading_column):
    workbook = openpyxl.load_workbook(file)
    worksheet = workbook.get_sheet_by_name(sheet)
    return worksheet.cell(row=reading_row, column=reading_column).value
